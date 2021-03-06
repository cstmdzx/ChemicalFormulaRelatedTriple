import xlrd
import xlwt
import re
import pdb
from openpyxl import load_workbook

excelFromRDF = load_workbook('ChemicalRelated.xlsx')
excelFromConcept = load_workbook('EquationRelatedReactant.xlsx')

sheet1FromRDF = excelFromRDF['Sheet1']
sheet1FromConcept = excelFromConcept['Sheet1']
sheet2FromConcept = excelFromConcept['Sheet2']

rowsFromRDF = sheet1FromRDF.rows
rowsFromConcept = sheet1FromConcept.rows

setFromRDF = set()

for eachRow in rowsFromRDF:
    line = [col.value for col in eachRow]
    if line[0] not in setFromRDF:
        setFromRDF.add(line[0])
    else:
        continue

setUnfinished = set()
for eachRow in rowsFromConcept:
    line = [col.value for col in eachRow]
    if line[0] not in setFromRDF:
        setUnfinished.add(line[0])
    else:
        continue

# print sheetUnfinished
intLine = 0
# print sheetUnfinished.cell(1, 1).value
patternEquationName = re.compile(u'([\u4e00-\u9fa5]+)\u4e0e([\u4e00-\u9fa5]+)\u7684\u53cd\u5e94')
for eachItem in setUnfinished:
    # sheetUnfinished.put_cell(intLine, 0, 1, eachItem, 0)
    match = patternEquationName.search(eachItem)
#    pdb.set_trace()
    if not match:
        print 'match error'
        print eachItem
        # continue
        strReactant1 = ''
        strReactant2 = ''
    else:
        strReactant1 = match.group(1)
        strReactant2 = match.group(2)
    intLine += 1
    sheet2FromConcept.cell(row = intLine, column = 1, value = eachItem)
    sheet2FromConcept.cell(row = intLine, column = 2, value = u'\u53cd\u5e94\u7269')
    sheet2FromConcept.cell(row = intLine, column = 3, value = strReactant1)
    # sheetUnfinished.put_cell(intLine, 0, 1, eachItem, 0)
    intLine += 1
    sheet2FromConcept.cell(row = intLine, column = 1, value = eachItem)
    sheet2FromConcept.cell(row = intLine, column = 2, value = u'\u53cd\u5e94\u7269')
    sheet2FromConcept.cell(row = intLine, column = 3, value = strReactant2)

excelFromConcept.save('EquationRelatedReactant.xlsx')



